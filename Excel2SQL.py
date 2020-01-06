import argparse
import os
import datetime

# importing for file selection in a directory
import glob
# importing for excel and sql work
from openpyxl import load_workbook
from openpyxl import Workbook

from os.path import join

# define and parse arguments
temp_arguments = argparse.ArgumentParser(description='This script will convert your excel sheets to sql code')

temp_arguments.add_argument("-i", "--input", required=True,
                            help="Name of the file or the directory of files")
temp_arguments.add_argument("-d", "--directory", default="Converted Code",
                            help="The directory where the converted files are to be stored. Type a folder name for "
                                 "storage in current folder or a file path for elsewhere")
passed_arguments = vars(temp_arguments.parse_args())

file_set = []  # stores .xlsx files to be converted

# set up an iterator of file paths or a just store a single file path depending on input
if ".xlsx" in passed_arguments["input"]:
    file_set.append(passed_arguments["input"])
else:
    file_set = glob.glob(join(passed_arguments["input"], "*" + ".xlsx"))

# check if the directory was not passed or was not a path and set directory accordingly
if passed_arguments["directory"] == "Converted Code" or "/" not in passed_arguments["directory"] or "\\" not in \
        passed_arguments["directory"]:
    if ".xlsx" in passed_arguments["input"]:
        inputPath = os.path.dirname(os.path.realpath(passed_arguments["input"]))
    else:
        inputPath = passed_arguments["input"]
    passed_arguments["directory"] = join(inputPath,
                                         passed_arguments["directory"])
print(passed_arguments["directory"])
# make directory if it doesn't exist
if not os.path.exists(passed_arguments["directory"]):
    os.makedirs((passed_arguments["directory"]))

for sheetPath in file_set:
    print(sheetPath)

    wb = load_workbook(sheetPath)
    ws = wb.active

    filenameIndex = sheetPath.rfind("/") if sheetPath.rfind("/") > sheetPath.rfind("\\") else sheetPath.rfind("\\")
    filenameIndex = filenameIndex if filenameIndex > 0 else 0
    sqlFileDir = passed_arguments["directory"] + "\\" + sheetPath[filenameIndex:-4] + "sql"
    sql_file = open(sqlFileDir, "w+")

    sql_file.write("INSERT ALL\n")
    is_table = False
    is_header = False
    is_table_name = False

    table_name = ""
    table_header_list = []
    table_header = ""
    values_list = []

    column_count = 0
    for row in ws.iter_rows(min_row=1, max_col=None, max_row=ws.max_row):

        for cell in row:
            column_count += 1
            if cell.value == "~":
                is_table_name = True
                is_table = False
                table_header_list.clear()
                values_list.clear()
                break

            if not is_table:
                if cell.value == "TABLENAME":
                    is_table_name = False
                    break
                if cell.value is not None:
                    table_name = cell.value

            else:
                if is_header:
                    if cell.value == "TABLEHEADER":
                        break
                    table_header_list.append(cell.value)
                else:
                    if cell.value is None:
                        values_list.append("NULL")
                    elif isinstance(cell.value, str):
                        try:
                            value = int(cell.value)
                            values_list.append(cell.value)
                        except ValueError:
                            values_list.append("\'" + cell.value + "\'")
                    elif isinstance(cell.value, datetime.datetime):
                        values_list.append("TO_DATE(\'" + str(cell.value) + "\',\'yyyy/mm/dd hh24:mi:ss\')")
                    elif isinstance(cell.value, int) or isinstance(cell.value, float):
                        values_list.append(int(cell.value))

                    if column_count == len(table_header_list):
                        break

        if not is_table_name:
            column_count = 0
            if not is_table:
                is_table = True
                is_header = True

            elif is_table and is_header:
                table_header = '(' + ','.join(table_header_list) + ')'
                is_header = False

            elif is_table and not is_header:
                sql_file.write("\tINTO " + table_name + " " + table_header + " VALUES " + "(" + ", ".join(
                    map(str, values_list)) + ")\n")
                values_list.clear()

    sql_file.write("SELECT * FROM dual;")

    sql_file.close()
